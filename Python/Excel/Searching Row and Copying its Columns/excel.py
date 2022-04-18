#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&#
# Created by: Mert PEHLIVANCIK
# Date: 18/04/2022 4:55 PM
# Github account: https://github.com/mertpehlivancik
# Copyright © 2022 Mert PEHLIVANCIK. All rights reserved.
#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&#
# If you want to make excel file, you should use this command
    ## pip install pyinstaller
# After run this command
    ## pyinstaller --onefile pythonScriptName.py
# Explanation of script:
    ## This script is used for that searches and finds the searched keyword in the given column and copies the entire row to the newly created excel file. It performs this operation for all excel files in the directory which is same with directory of script. 
from logging import lastResort
import sys
from winreg import ConnectRegistry
from openpyxl import load_workbook
import openpyxl
import re
import pathlib
from datetime import date

if __name__ == '__main__':
    print("Github account: https://github.com/mertpehlivancik")
    print("Copyright © 2022 Mert PEHLIVANCIK. All rights reserved.")
    usedSheet = input("Enter the name of the sheet you want to search(Please give input paying attention case sensitively): ")
    whichColumn = int(input("Enter the column you want to search(Please give an integer as a column. For example: A = 1): "))
    searchItem = input("Enter the name of the item you want to search: ")
    newSheet = input("Enter the name of the new sheet you want to create: ")
    newFilename = input("Enter the name of the new filename you want to create: ")
    searchItem = searchItem.lower()
    lastRow = 2
    nWorkbook = openpyxl.Workbook()
    nDetaySheet = nWorkbook.create_sheet(newSheet)
    currentDirectory = pathlib.Path('.')
    currentPattern = "*.xlsx"
    listExcelFiles = []
    isSheetFound = False
    for currentFile in currentDirectory.glob(currentPattern):
        listExcelFiles.append(currentFile)
    
    currentPath = ""
    for cf in listExcelFiles:
        currentPath = cf
        workbook = load_workbook(filename=currentPath)
        for wbs in workbook.sheetnames:
            if wbs == usedSheet:
                detaySheet = workbook[usedSheet]
                SearchedRow = []
                isSheetFound = True
                for i in range(1, detaySheet.max_row+1):
                    foundcell = detaySheet.cell(row=i, column=whichColumn).value
                    foundcell = str(foundcell)
                    foundcell = foundcell.lower()
                    if foundcell.find(searchItem) != -1:
                        SearchedRow.append(i)
                for j in SearchedRow:   
                    for k in range(1, detaySheet.max_column+1):
                        nDetaySheet.cell(row=lastRow, column=k).value = detaySheet.cell(row=j, column=k).value
                    lastRow += 1
      
    if bool(isSheetFound) == False:
        print("Sheet not found!")
        sys.exit()  
    today = date.today()
    Lastfilename = newFilename + today.strftime("%d%m%Y") + ".xlsx"
    nWorkbook.save(Lastfilename)
    nWorkbook.close()

        

