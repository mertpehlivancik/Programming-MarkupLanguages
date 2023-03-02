#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&#
# Created by: Mert PEHLIVANCIK
# Date: 07/02/2023 4:55 PM
# Github account: https://github.com/mertpehlivancik
# Copyright © 2023 Mert PEHLIVANCIK. All rights reserved.
#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&#
# If you want to make excel file, you should use this command
# pip install pyinstaller
# After run this command
# pyinstaller --onefile pythonScriptName.py
# Explanation of script:
from logging import lastResort
import sys
from winreg import ConnectRegistry
from openpyxl import load_workbook
import openpyxl
import re
import pathlib
from datetime import date

if __name__ == '__main__':
    print("Github account: https://github.com/mertpehlivancik")
    print("Copyright © 2023 Mert PEHLIVANCIK. All rights reserved.")
    print("KARABAĞLAR ARÇELİK BEKO YETKİLİ SERVİSİ")
    while True:
        secenek = input("Programı tamamlayıcı için kullanacaksanız 't', ek garanti için kullanacaksanız 'e' harfini tuşlayınız: ")
        if secenek == "t":
            ## Tamamlayıcı
            print("SATIŞ FİYATLARI LİSTESİ")
            fiyatFileName = input("Dosyanın adını giriniz(fiyat.xlsx): ")
            fiyatSheetName = input("Sheet adını giriniz(Tamamlayici_Urunler): ")
            stokColumn = int(input("Stok numaralarının bulunduğu kolon(Sayı olarak giriniz. Örn: B = 2): "))
            fiyatColumn = int(input("Ödenecek pay bedellerinin bulunduğu kolon(Sayı olarak giriniz. Örn: L = 12): "))
            print("TEKNİSYEN SATIŞ LİSTESİ")
            tamamlayıcıSatisFileName = input("Dosyasının adını giriniz(tamamlayiciSatis.xlsx): ")
            tamamlayiciSatisSheetName = input("Sheet adını giriniz(TamamlayiciUrunSatis): ")
            tamamlayiciSatisMiktarColumn = int(input("Satış miktarlarının kaydedildiği kolonu giriniz(Sayı olarak giriniz. Örn: G = 7): "))
            tamamlayiciSatisMiktarRow = int(input("Satış miktarlarının kaydedilmeye başladığı satırı giriniz(Sayı olarak giriniz. Örn: 6): "))
            tamamlayiciSatisBakiyeKayit = int(input("Sonuç hangi kolona kaydedilsin?(Sayı olarak giriniz. Örn: L = 12): "))
            baslik = input("Sonuçların başlığı ne olsun?: ")
            fiyatPath = fiyatFileName + ".xlsx"
            tamamlayiciPath = tamamlayıcıSatisFileName + ".xlsx"

            workbookFiyat = load_workbook(filename=fiyatPath)
            workbookSatis = load_workbook(filename=tamamlayiciPath)
            fiyatSheet = workbookFiyat[fiyatSheetName]
            satisSheet = workbookSatis[tamamlayiciSatisSheetName]
            satisSheet.delete_cols(idx = tamamlayiciSatisBakiyeKayit)
            satisSheet.insert_cols(idx = tamamlayiciSatisBakiyeKayit)
            satisSheet.cell(row=tamamlayiciSatisMiktarRow-1, column=tamamlayiciSatisBakiyeKayit).value = baslik
            dict = {}
            listStokNo = []

            for i in range(1, fiyatSheet.max_row+1):
                stokNo = fiyatSheet.cell(row=i, column=stokColumn).value
                fiyat = fiyatSheet.cell(row=i, column=fiyatColumn).value
                stokNo = str(stokNo)
                fiyat = str(fiyat)
                if stokNo == "None":
                    continue
                dict[stokNo] = fiyat
                listStokNo.append(stokNo)
                
            for s in range(tamamlayiciSatisMiktarRow, satisSheet.max_row+1):
                satis = satisSheet.cell(row=s, column=tamamlayiciSatisMiktarColumn).value
                satis = str(satis)
                 
                bulunanIndex = []
                bulunanStok = []
                bulunanAdet = []
                adet = ""

                for m in range(0, len(listStokNo)):
                    searching = str(listStokNo[m])
                    index = satis.find(searching)
                    if(index != -1):
                        bulunanIndex.append(index)
                        bulunanStok.append(searching)
                        index = index + len(searching)
                        index = index + 1
                        while satis[index] != " ":
                            adet = adet + satis[index]
                            index = index + 1
                        bulunanAdet.append(adet)

                        nAdet = float(adet)
                        nSearchingFiyat = float(dict[searching])
                        if satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value == None:
                            satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value = nAdet * nSearchingFiyat
                            #print(satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value)
                            #print(nAdet ,nSearchingFiyat,nAdet * nSearchingFiyat)
                            adet = ""
                        else:
                            satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value = satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value + nAdet * nSearchingFiyat
                            #print(satisSheet.cell(row=s, column=tamamlayiciSatisBakiyeKayit).value)
                            #print(nAdet ,nSearchingFiyat, satisSheet.cell(row=s, column=12).value + nAdet * nSearchingFiyat)
                            adet = ""
            workbookSatis.save(filename=tamamlayiciPath)
            break
        elif secenek == "e":
            print("SATIŞ FİYATLARI LİSTESİ")
            fiyatFileName = input("Dosyanın adını giriniz(fiyat.xlsx): ")
            fiyatSheetName = input("Sheet adını giriniz(Sayfa1): ")
            fiyatHizmetKodu = int(input("Hizmet kodlarının olduğu kolon numarası(B = 2): "))
            fiyatTeknisyenPayı = int(input("Ödenecek pay bedellerinin bulunduğu kolon(F = 6): "))
            print("TEKNİSYEN SATIŞ LİSTESİ")
            EkGarantiSatisFileName = input("Dosyanın adını giriniz(EkGarantiSatis.xlsx): ")
            EkGarantiSatisSheetName = input("Sheet adını giriniz(EkGarantiSatis): ")
            EkGarantiSatisKayıtColumn = int(input("Satışların kaydedildiği kolonu giriniz(Sayı olarak giriniz. Örn: F = 6): "))
            EkGarantiSatisRow = int(input("Satışların başladığı satır numarası(6): "))
            EkGarantiBakiyeKayit = int(input("Sonuç hangi kolona kaydedilsin?(Sayı olarak giriniz. Örn: K = 11): "))
            baslik = input("Sonuçların başlığı ne olsun?: ")
            
            fiyatPath = fiyatFileName + ".xlsx"
            tamamlayiciPath = EkGarantiSatisFileName + ".xlsx"

            workbookFiyat = load_workbook(filename=fiyatPath)
            workbookSatis = load_workbook(filename=tamamlayiciPath)
            fiyatSheet = workbookFiyat[fiyatSheetName]
            satisSheet = workbookSatis[EkGarantiSatisSheetName]
            satisSheet.delete_cols(idx = EkGarantiBakiyeKayit)
            satisSheet.insert_cols(idx = EkGarantiBakiyeKayit)
            satisSheet.cell(row=EkGarantiSatisRow-1, column=EkGarantiBakiyeKayit).value = baslik
            dict = {}
            listStokNo = []

            for i in range(1, fiyatSheet.max_row+1):
                stokNo = fiyatSheet.cell(row=i, column=fiyatHizmetKodu).value
                fiyat = fiyatSheet.cell(row=i, column=fiyatTeknisyenPayı).value
                stokNo = str(stokNo)
                fiyat = str(fiyat)
                if stokNo == "None":
                    continue
                dict[stokNo] = fiyat
                listStokNo.append(stokNo)

            for s in range(EkGarantiSatisRow, satisSheet.max_row+1):
                satis = satisSheet.cell(row=s, column=EkGarantiSatisKayıtColumn).value
                satis = str(satis)

                bulunanIndex = []
                bulunanStok = []
                bulunanAdet = []
                adet = ""

                for m in range(0, len(listStokNo)):
                    searching = str(listStokNo[m])
                    index = satis.find(searching)
                    if(index != -1):
                        nSearchingFiyat = float(dict[searching])
                        satisSheet.cell(row=s, column=EkGarantiBakiyeKayit).value = nSearchingFiyat

            workbookSatis.save(filename=tamamlayiciPath)
            break
        else:
            print("Yanlış seçenek girdiniz. Lütfen tekrar deneyiniz.")
            continue
    print("İşlem başarıyla tamamlanmıştır.")
    input("Kapatmak için bir tuşa basınız...")
    print("Teşekkürler :)")
    