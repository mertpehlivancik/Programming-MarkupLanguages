from openpyxl import load_workbook
import openpyxl 

if __name__ == '__main__':
    workbook = load_workbook(filename="malzeme.xlsx")    
    sheet = workbook.active
    c = sheet.cell(row= 1 , column = 19)
    print(c.value) 
    
    count = 0
    list_malzeme = []
    list_sayi = []
    
    for x in range(2,18508):
        c = sheet.cell(row= x , column = 19)
        c = str(c.value)
        list_malzeme.append(c)
    
    mylist = list(dict.fromkeys(list_malzeme))
#     print(mylist)
    
    for y in mylist:
        for z in range(2,18508):
            d = sheet.cell(row= z , column = 19)
            d = str(d.value)
            y = str(y)
            if y == d:
                e = sheet.cell(row= z , column = 58)
                count = count + int(e.value)
        list_sayi.append(count)
        count = 0
       
       
    wb = openpyxl.Workbook() 
    sheet2 = wb.active    
    
    
    
    g = sheet2.cell(row = 1, column = 1)
    g.value = "Malzeme Stok Kodu"
    f = sheet2.cell(row = 1, column = 2)
    f.value = "Toplam Adet"
    
    count = 2 
    for z in mylist:
        g = sheet2.cell(row = count, column = 1)
        g.value = z
        count = count + 1
    
    count = 2
    for m in list_sayi:
        g = sheet2.cell(row = count, column = 2)
        g.value = m
        count = count + 1
    
    wb.save(filename="Malzeme_Total.xlsx") 
    print(list_malzeme)
    print(mylist)
    print(list_sayi)

        